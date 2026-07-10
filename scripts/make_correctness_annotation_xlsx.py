import json, os, random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

SP = Path(os.environ["TEMP"]) / "claude" / "C--Users-Administrator-paft-paper" / "f9e848c1-78b3-45ba-91f3-9e7432b08b84" / "scratchpad"
OUT = Path(r"C:\Users\Administrator\paft-annotation")
OUT.mkdir(exist_ok=True)

rows = json.loads((SP / "annotation_rows.json").read_text(encoding="utf-8"))
rng = random.Random(42)
rng.shuffle(rows)
for i, r in enumerate(rows, 1):
    r["pid"] = f"P{i:03d}"

HEAD_FILL = PatternFill("solid", start_color="1F4E79")
HEAD_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
CODE_FONT = Font(name="Consolas", size=9)
TXT_FONT = Font(name="Microsoft YaHei", size=10)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

def clip(s, n=30000):
    s = s or ""
    return s if len(s) <= n else s[:n] + "\n...[截断]"

# ---------- main blinded workbook ----------
wb = Workbook()

ws0 = wb.active
ws0.title = "标注说明"
ws0.column_dimensions["A"].width = 110
guide = [
    ("PAFT 补丁正确性人工标注", True),
    ("", False),
    ("任务：判断每个候选补丁是否真正修复了缺陷（语义正确），而不仅仅是通过了测试套件。", False),
    ("", False),
    ("背景：表中每一行是 Defects4J 某个 bug 的一个候选补丁。该补丁已通过基准的全部测试（plausible），", False),
    ("但测试套件可能不完整，通过测试不代表语义正确。请以“开发者修复（参考）”列为语义基准进行判断。", False),
    ("", False),
    ("标注选项（在“标注结果”列的下拉菜单中选择）：", True),
    ("  正确 —— 候选补丁与开发者修复语义等价，或以其他方式正确修复了缺陷（允许实现方式不同）。", False),
    ("  疑似过拟合 —— 候选补丁能通过现有测试，但与开发者修复语义不一致，", False),
    ("      例如仅针对触发测试打补丁、遗漏边界情况、或改变了规定之外的行为。", False),
    ("  不确定 —— 依据给出的信息无法判断（请在“备注”列简述原因）。", False),
    ("", False),
    ("注意事项：", True),
    ("  1. 候选补丁的生成方法已匿名化，请独立判断每一行，不要猜测其来源。", False),
    ("  2. 只需判断语义正确性，不要考虑代码风格、注释或改动大小。", False),
    ("  3. 每行独立成题；相同 Bug 会出现多行（不同候选补丁），请分别判断。", False),
    ("  4. 如需查看完整上下文，可按“Bug”列编号在 Defects4J 中检索。", False),
    ("", False),
    (f"共 {len(rows)} 行。祝标注顺利！", False),
]
for i, (text, bold) in enumerate(guide, 1):
    c = ws0.cell(row=i, column=1, value=text)
    c.font = Font(name="Microsoft YaHei", size=12 if i == 1 else 10, bold=bold)
    c.alignment = Alignment(wrap_text=True, vertical="top")

ws = wb.create_sheet("标注表")
headers = ["编号", "Bug", "问题标题", "缺陷代码", "候选补丁", "开发者修复（参考）", "标注结果", "备注"]
widths = [8, 16, 30, 60, 60, 60, 14, 24]
for j, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=j, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN
    ws.column_dimensions[chr(64 + j)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

for i, r in enumerate(rows, 2):
    vals = [r["pid"], r["bug"], clip(r["issue_title"], 500), clip(r["buggy"]), clip(r["patch"]), clip(r["dev_fix"]), "", ""]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=j, value=v)
        c.border = THIN
        c.alignment = WRAP_TOP
        c.font = CODE_FONT if j in (4, 5, 6) else TXT_FONT
    ws.row_dimensions[i].height = 160

dv = DataValidation(type="list", formula1='"正确,疑似过拟合,不确定"', allow_blank=True, showDropDown=False)
dv.error = "请从下拉列表中选择：正确 / 疑似过拟合 / 不确定"
dv.errorTitle = "无效输入"
ws.add_data_validation(dv)
dv.add(f"G2:G{len(rows) + 1}")

wb.save(OUT / "PAFT-补丁正确性标注表.xlsx")

# ---------- key workbook (method mapping, keep private) ----------
kb = Workbook()
ks = kb.active
ks.title = "对照密钥"
for j, h in enumerate(["编号", "Bug", "方法"], 1):
    c = ks.cell(row=1, column=j, value=h)
    c.font = HEAD_FONT
    c.fill = HEAD_FILL
ks.column_dimensions["A"].width = 10
ks.column_dimensions["B"].width = 18
ks.column_dimensions["C"].width = 12
for i, r in enumerate(rows, 2):
    ks.cell(row=i, column=1, value=r["pid"]).font = TXT_FONT
    ks.cell(row=i, column=2, value=r["bug"]).font = TXT_FONT
    ks.cell(row=i, column=3, value=r["method"]).font = TXT_FONT
ks.freeze_panes = "A2"
note = ks.cell(row=1, column=5, value="注意：本文件为方法对照密钥，请在标注完成前不要发给标注者。")
note.font = Font(name="Microsoft YaHei", bold=True, color="C00000")
kb.save(OUT / "PAFT-标注方法对照密钥.xlsx")

from collections import Counter
print("rows:", len(rows), dict(Counter(r["method"] for r in rows)))
print("saved:", OUT / "PAFT-补丁正确性标注表.xlsx")
print("saved:", OUT / "PAFT-标注方法对照密钥.xlsx")
